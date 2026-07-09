import csv
import io
import json
import re
from datetime import datetime, time, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_, select, nulls_last
from sqlalchemy.orm import Session

from .models import ScheduleEntry, Student
from .schemas import SEM_FIELDS, ImportResponse, StatsResponse


DEFAULT_EVENT_COLOR = "#3bd688"
DEFAULT_EVENT_TYPE = "Placement Drive"
SORTABLE_FIELDS = {
    "name": Student.name,
    "usn": Student.usn,
    "cgpa": Student.cgpa,
    "tenth_pct": Student.tenth_pct,
    "twelfth_pct": Student.twelfth_pct,
    "active_backlogs": Student.active_backlogs,
    "mobile": Student.mobile,
    "personal_email": Student.personal_email,
    "college_email": Student.college_email,
    "placement_status": Student.placement_status,
}


def normalize_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_header(value) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def normalize_usn(usn: str) -> str:
    value = normalize_text(usn)
    return value.upper() if value else ""


def parse_float(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_int(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "placed", "active"}:
        return True
    if text in {"0", "false", "no", "n", "none", ""}:
        return False
    return bool(text)


def clean_student_payload(payload: dict) -> dict:
    cleaned = dict(payload)
    placement_status = parse_bool(cleaned.get("placement_status", cleaned.get("is_placed")))
    company_name = normalize_text(cleaned.get("company_name", cleaned.get("placed_company"))) if placement_status else None

    cleaned["name"] = normalize_text(cleaned.get("name")) or ""
    cleaned["usn"] = normalize_usn(cleaned.get("usn"))
    cleaned["program"] = normalize_text(cleaned.get("program")) or "ECE"
    cleaned["mobile"] = normalize_text(cleaned.get("mobile"))
    cleaned["personal_email"] = normalize_text(cleaned.get("personal_email"))
    cleaned["college_email"] = normalize_text(cleaned.get("college_email"))
    cleaned["tenth_pct"] = parse_float(cleaned.get("tenth_pct"))
    cleaned["twelfth_pct"] = parse_float(cleaned.get("twelfth_pct"))
    for field in SEM_FIELDS:
        cleaned[field] = parse_float(cleaned.get(field))
    cleaned["cgpa"] = parse_float(cleaned.get("cgpa"))
    cleaned["placement_status"] = placement_status
    cleaned["company_name"] = company_name
    cleaned["is_placed"] = placement_status
    cleaned["placed_company"] = company_name
    cleaned["active_backlogs"] = parse_int(cleaned.get("active_backlogs")) or 0
    cleaned["backlog_current"] = "YES" if cleaned["active_backlogs"] > 0 else "NO"
    return cleaned


def apply_student(student: Student, payload: dict) -> Student:
    cleaned = clean_student_payload(payload)
    for key, value in cleaned.items():
        setattr(student, key, value)
    return student


def list_students(
    db: Session,
    query: str = "",
    backlog_filter: str = "all",
    placement_filter: str = "all",
    sort_key: str = "name",
    sort_dir: str = "asc",
):
    stmt = select(Student)

    if query:
        like = f"%{query.strip()}%"
        stmt = stmt.where(
            or_(
                Student.name.ilike(like),
                Student.usn.ilike(like),
                Student.company_name.ilike(like),
            )
        )

    if backlog_filter == "ok":
        stmt = stmt.where(Student.active_backlogs == 0)
    elif backlog_filter == "warn":
        stmt = stmt.where(Student.active_backlogs > 0)

    if placement_filter == "placed":
        stmt = stmt.where(Student.placement_status.is_(True))
    elif placement_filter == "not-placed":
        stmt = stmt.where(Student.placement_status.is_(False))

    order_field = SORTABLE_FIELDS.get(sort_key, Student.name)

    NUMERIC_FIELDS = {
        "cgpa",
        "tenth_pct",
        "twelfth_pct",
        "active_backlogs",
    }

    if sort_key in NUMERIC_FIELDS:
        if sort_dir == "desc":
            stmt = stmt.order_by(
                nulls_last(order_field.desc()),
                Student.name.asc(),
            )
        else:
            stmt = stmt.order_by(
                nulls_last(order_field.asc()),
                Student.name.asc(),
            )
    else:
        if sort_dir == "desc":
            stmt = stmt.order_by(
                order_field.desc(),
                Student.name.asc(),
            )
        else:
            stmt = stmt.order_by(
                order_field.asc(),
                Student.name.asc(),
            )

    return db.execute(stmt).scalars().all()


def get_stats(db: Session) -> StatsResponse:
    total = db.scalar(select(func.count()).select_from(Student)) or 0
    average = db.scalar(select(func.avg(Student.cgpa)).where(Student.cgpa.is_not(None))) or 0
    top = db.scalar(select(func.max(Student.cgpa)).where(Student.cgpa.is_not(None))) or 0
    backlogs = db.scalar(select(func.count()).select_from(Student).where(Student.active_backlogs > 0)) or 0
    zero_backlog = db.scalar(select(func.count()).select_from(Student).where(Student.active_backlogs == 0)) or 0
    placed_students = db.scalar(select(func.count()).select_from(Student).where(Student.placement_status.is_(True))) or 0
    not_placed_students = db.scalar(select(func.count()).select_from(Student).where(Student.placement_status.is_(False))) or 0
    return StatsResponse(
        total=total,
        average_cgpa=round(float(average), 2) if average else 0,
        active_backlogs=backlogs,
        top_cgpa=round(float(top), 2) if top else 0,
        zero_backlog=zero_backlog,
        placed_students=placed_students,
        not_placed_students=not_placed_students,
    )


def find_student_by_usn(db: Session, usn: str) -> Student | None:
    return db.scalar(select(Student).where(Student.usn == normalize_usn(usn)))


def export_csv(students: list[Student]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    headers = [
        "id", "name", "usn", "program", "mobile", "personal_email", "college_email",
        "tenth_pct", "twelfth_pct", "sem1", "sem2", "sem3", "sem4", "sem5", "sem6", "sem7", "sem8",
        "cgpa", "placement_status", "company_name", "backlog_current", "active_backlogs",
    ]
    writer.writerow(headers)
    for student in students:
        writer.writerow([getattr(student, field) for field in headers])
    return output.getvalue()


def export_xlsx(students: list[Student]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = [
        "name", "usn", "program", "mobile", "personal_email", "college_email",
        "tenth_pct", "twelfth_pct", "sem1", "sem2", "sem3", "sem4", "sem5", "sem6", "sem7", "sem8",
        "cgpa", "placement_status", "company_name", "active_backlogs",
    ]
    ws.append(headers)
    for student in students:
        ws.append([getattr(student, field) for field in headers])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def import_students_from_workbook(db: Session, file_bytes: bytes) -> ImportResponse:
    workbook = load_workbook(io.BytesIO(file_bytes))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ImportResponse(created=0, updated=0, ignored=0)

    headers = [normalize_text(value) for value in rows[0]]
    created = updated = ignored = 0
    seen_usns: set[str] = set()

    header_aliases = {
        "name": "name",
        "usn": "usn",
        "program": "program",
        "mobile": "mobile",
        "personalemail": "personal_email",
        "collegeemail": "college_email",
        "tenthpct": "tenth_pct",
        "twelfthpct": "twelfth_pct",
        "sem1": "sem1",
        "semester1": "sem1",
        "sem01": "sem1",
        "sem2": "sem2",
        "semester2": "sem2",
        "sem02": "sem2",
        "sem3": "sem3",
        "semester3": "sem3",
        "sem03": "sem3",
        "sem4": "sem4",
        "semester4": "sem4",
        "sem04": "sem4",
        "sem5": "sem5",
        "semester5": "sem5",
        "sem05": "sem5",
        "sem6": "sem6",
        "semester6": "sem6",
        "sem06": "sem6",
        "sem7": "sem7",
        "semester7": "sem7",
        "sem07": "sem7",
        "sem8": "sem8",
        "semester8": "sem8",
        "sem08": "sem8",
        "cgpa": "cgpa",
        "cumulativecgpa": "cgpa",
        "overallcgpa": "cgpa",
        "cgpadetails": "cgpa",
        "finalcgpa": "cgpa",
        "isplaced": "placement_status",
        "placed": "placement_status",
        "placementstatus": "placement_status",
        "placedcompany": "company_name",
        "company": "company_name",
        "companyname": "company_name",
        "activebacklogs": "active_backlogs",
        "activebacklog": "active_backlogs",
        "backlogs": "active_backlogs",
        "backlogcurrent": "backlog_current",
        "backlog": "backlog_current",
        "personalemailid": "personal_email",
        "collegeemailid": "college_email",
        "10thpercentage": "tenth_pct",
        "10thpct": "tenth_pct",
        "10th": "tenth_pct",
        "12thpercentage": "twelfth_pct",
        "12thpct": "twelfth_pct",
        "12th": "twelfth_pct",
        "1stsemsgpa": "sem1",
        "2ndsemsgpa": "sem2",
        "3rdsemsgpa": "sem3",
        "4thsemsgpa": "sem4",
        "5thsemsgpa": "sem5",
        "6thsemsgpa": "sem6",
        "7thsemsgpa": "sem7",
        "8thsemsgpa": "sem8",
        "cummulativecgpa": "cgpa",
        "currentbacklogsubjectsyesno": "backlog_current",
        "currentbacklogsubjects": "backlog_current",
        "numberofactivebacklogs": "active_backlogs",
    }

    for raw in rows[1:]:
        if not any(value not in (None, "") for value in raw):
            continue
        row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw))) if headers[i]}
        normalized_row = {}
        for header, value in row.items():
            canonical = header_aliases.get(normalize_header(header))
            if canonical:
                normalized_row[canonical] = value

        usn = normalize_usn(normalized_row.get("usn"))
        name = normalize_text(normalized_row.get("name"))
        if not usn or not name:
            ignored += 1
            continue
        if usn in seen_usns:
            ignored += 1
            continue
        seen_usns.add(usn)

        payload = {
            "name": name,
            "usn": usn,
            "program": normalize_text(normalized_row.get("program")) or "ECE",
            "mobile": normalized_row.get("mobile"),
            "personal_email": normalized_row.get("personal_email"),
            "college_email": normalized_row.get("college_email"),
            "tenth_pct": normalized_row.get("tenth_pct"),
            "twelfth_pct": normalized_row.get("twelfth_pct"),
            "sem1": normalized_row.get("sem1"),
            "sem2": normalized_row.get("sem2"),
            "sem3": normalized_row.get("sem3"),
            "sem4": normalized_row.get("sem4"),
            "sem5": normalized_row.get("sem5"),
            "sem6": normalized_row.get("sem6"),
            "sem7": normalized_row.get("sem7"),
            "sem8": normalized_row.get("sem8"),
            "cgpa": normalized_row.get("cgpa"),
            "placement_status": normalized_row.get("placement_status"),
            "company_name": normalized_row.get("company_name"),
            "active_backlogs": normalized_row.get("active_backlogs"),
        }
        existing = find_student_by_usn(db, usn)
        if existing:
            apply_student(existing, payload)
            updated += 1
        else:
            db.add(apply_student(Student(), payload))
            created += 1

    db.commit()
    return ImportResponse(created=created, updated=updated, ignored=ignored)


def seed_from_legacy_html(db: Session):
    if db.scalar(select(func.count()).select_from(Student)):
        return

    html_path = Path(__file__).resolve().parents[2] / "ece-placement-portal.html"
    if not html_path.exists():
        return

    text = html_path.read_text()
    match = re.search(r"const STUDENTS = (\[.*?\]);\n\nconst ADMIN_CODE", text, re.S)
    if not match:
        return

    records = json.loads(match.group(1))
    for record in records:
        record.setdefault("placement_status", record.get("is_placed", False))
        record.setdefault("company_name", record.get("placed_company"))
        db.add(apply_student(Student(), record))
    db.commit()


def _default_end_time(start_at: datetime) -> datetime:
    return start_at + timedelta(hours=1)


def clean_schedule_payload(payload: dict, existing: ScheduleEntry | None = None) -> dict:
    cleaned = dict(payload)
    start_at = cleaned.get("start_at") or getattr(existing, "start_at", None)
    end_at = cleaned.get("end_at") or getattr(existing, "end_at", None)
    drive_date = cleaned.get("drive_date") or getattr(existing, "drive_date", None)

    if start_at is None and drive_date is not None:
        start_at = datetime.combine(drive_date, time(hour=9))
    if drive_date is None and start_at is not None:
        drive_date = start_at.date()
    if end_at is None and start_at is not None:
        end_at = _default_end_time(start_at)

    location = normalize_text(cleaned.get("location", cleaned.get("venue")))
    if location is None:
        location = getattr(existing, "location", None) or getattr(existing, "venue", None)
    company_name = normalize_text(cleaned.get("company_name"))
    if company_name is None:
        company_name = getattr(existing, "company_name", None)
    title = normalize_text(cleaned.get("title")) or getattr(existing, "title", None) or company_name or DEFAULT_EVENT_TYPE
    if company_name is None:
        company_name = title

    return {
        "title": title,
        "company_name": company_name,
        "drive_date": drive_date,
        "start_at": start_at,
        "end_at": end_at,
        "event_type": normalize_text(cleaned.get("event_type")) or getattr(existing, "event_type", None) or DEFAULT_EVENT_TYPE,
        "color": normalize_text(cleaned.get("color")) or getattr(existing, "color", None) or DEFAULT_EVENT_COLOR,
        "description": normalize_text(cleaned.get("description")) if "description" in cleaned else getattr(existing, "description", None),
        "venue": location,
        "location": location,
    }


def apply_schedule_entry(entry: ScheduleEntry, payload: dict) -> ScheduleEntry:
    cleaned = clean_schedule_payload(payload, existing=entry)
    for key, value in cleaned.items():
        setattr(entry, key, value)
    return entry


def list_schedule_entries(db: Session, search: str = "", company: str = "", event_type: str = ""):
    stmt = select(ScheduleEntry)
    if search:
        like = f"%{search.strip()}%"
        stmt = stmt.where(
            or_(
                ScheduleEntry.title.ilike(like),
                ScheduleEntry.company_name.ilike(like),
                ScheduleEntry.description.ilike(like),
                ScheduleEntry.location.ilike(like),
            )
        )
    if company:
        stmt = stmt.where(ScheduleEntry.company_name.ilike(f"%{company.strip()}%"))
    if event_type:
        stmt = stmt.where(ScheduleEntry.event_type == event_type)
    return db.execute(
        stmt.order_by(ScheduleEntry.start_at.asc(), ScheduleEntry.company_name.asc(), ScheduleEntry.title.asc())
    ).scalars().all()